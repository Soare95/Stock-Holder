from selenium import webdriver
from openpyxl import Workbook, load_workbook
import schedule
import time

def verify():
	driver = webdriver.Chrome(r'D:\Applications\Programs\Automation\chromedriver.exe')
	driver.get('https://finance.yahoo.com/quote/BTC-USD/')
	driver.implicitly_wait(10)

	accept_all = driver.find_element_by_xpath('//*[@id="consent-page"]/div/div/div/form/div[2]/div[2]/button').click()

	price = driver.find_element_by_xpath('//*[@id="quote-header-info"]/div[3]/div[1]/div/span[1]').text
	moving = driver.find_element_by_xpath('//*[@id="quote-header-info"]/div[3]/div[1]/div/span[2]').text

	myFileName=r'D:\Applications\Programs\Stocks-Holder\data-base.xlsx'
	wb = load_workbook(filename=myFileName)
	ws = wb.active
	ws['A1'] = 'Price'
	ws['B1'] = 'Move'
	ws.append([price, moving])
	wb.save("data-base.xlsx")

	driver.quit()

schedule.every().seconds.do(verify)

while True:
    schedule.run_pending()
    time.sleep(1)